from todoist_api_python.api import TodoistAPI
import os
from sys import argv
from dotenv import load_dotenv
from requests import post
from pyautogui import hotkey as hk
import pyautogui
from tabulate import tabulate

pyautogui.FAILSAFE = False

args = argv

if len(args) == 2:
    import pyautogui as pg
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    while True:
        try:
            fw = pg.getActiveWindow()
            if 'Morning-Routine' in fw.title:
                fw.activate()
                break
        except:
            pass
    workbook = load_workbook(
        r'C:\Users\Aarjav\Documents\Productivity Tracker.xlsx')
    worksheet = workbook.active

    # delete all cells
    worksheet.delete_cols(1, worksheet.max_column)
    worksheet.delete_rows(1, worksheet.max_row)

    headers = ["Task", "Estimate", "Time Taken", "Efficiency"]
    bold_font = Font(bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = bold_font

    # set up the AVERAGE row with custom cell style
    average_row = ["MEDIAN", None, None, "=ROUND(MEDIAN(D2:D2), 0)"]

    for col_num, value in enumerate(average_row, start=1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = value
        cell.style = "Good"
        cell.font = bold_font

    # center all cells in the worksheet
    for row in worksheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # save the workbook
    workbook.save(r'C:\Users\Aarjav\Documents\Productivity Tracker.xlsx')

hk('win', 'up')

load_dotenv(
    dotenv_path=r'C:\Users\Aarjav\Documents\Automation-Programs\Todoist-Hyperscheduler\todoist.env')
load_dotenv(
    r'C:\Users\Aarjav\Documents\Automation-Programs\Habitica\habitica.env')

key = os.environ.get('TODOIST_API_TOKEN')
api = TodoistAPI(key)

projects = {
    int(i.id): i.name for i in api.get_projects() if 'Dummy' not in i.name
}

study_tasks = api.get_tasks(label='study')
study_tasks.sort(key=lambda x: projects[int(x.project_id)])
tasks_names = [[i+1, task.content.split('](')[0].replace(
    '[', ''), projects[int(study_tasks[i].project_id)]] for i, task in enumerate(study_tasks)]

# Sort tasks_names by the third item in each item
tasks_names.sort(key=lambda x: x[2])

interest_tasks = api.get_tasks(label='interests')
tasks_names.append(['', '', ''])
tasks_names.extend(list([i+1, task.content.split('](')[0].replace(
    '[', ''), projects[int(interest_tasks[i-len(study_tasks)].project_id)]] for i, task in enumerate(interest_tasks, len(study_tasks))))
print(tabulate(tasks_names, headers=[
      '#', 'Task', 'Project']))
print()

# Ask user which task they want to choose as next
next = tasks_names[int(input('Enter task number: '))-1][1]

# Estimate how long it will take to complete the task
estimation = int(input('Estimate duration (minutes): '))

KEY = os.getenv('KEY')
USER = os.getenv('ID')
USER_CLIENT = f'{USER}-Testing'

# Post to https://habitica.com/api/v3/tasks/user
habitica_new_task = post('https://habitica.com/api/v3/tasks/user', headers={
    'x-api-user': USER,
    'x-api-key': KEY,
    'x-client': USER_CLIENT,
    'Content-Type': 'application/json'
}, json={
    'text': next,
    'type': 'todo',
    'notes': f'Estimated duration: {estimation} minutes',
})

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Load the workbook
workbook = load_workbook(
    r'C:\Users\Aarjav\Documents\Productivity Tracker.xlsx')

# Select the active worksheet
worksheet = workbook.active

# Insert a new row after the header

if '](' in next:
    next = next.split('](')[0].replace('[', '')

for row in reversed(range(1, worksheet.max_row + 1)):
    # Check if the row has any text
    if any(worksheet.cell(row=row, column=col).value for col in range(1, worksheet.max_column + 1)):
        # Found the last row with text
        new_row = row-1
        break

worksheet.insert_rows(new_row)

# Set the values for the new row
worksheet.cell(row=new_row, column=1).value, worksheet.cell(
    row=new_row, column=2).value, worksheet.cell(
    row=new_row, column=4).value = next, estimation, f'=IFERROR(ROUND((B{new_row}/C{new_row})*5,2),"")'

for i in range(1, 5):
    worksheet.cell(row=new_row, column=i).alignment = Alignment(
        horizontal='center', vertical='center')

# Save the changes to the workbook
workbook.save(r'C:\Users\Aarjav\Documents\Productivity Tracker.xlsx')
